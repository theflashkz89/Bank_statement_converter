"""
Excel导出模块 - 将标准化后的数据导出为Excel文件
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, Any
import logging
from pathlib import Path


logger = logging.getLogger(__name__)


class FileLockedError(Exception):
    """文件被占用异常"""
    pass


def export_to_excel(df: pd.DataFrame, summary: Dict[str, Any], output_path: str) -> str:
    """
    导出数据到Excel文件
    
    参数:
        df: 标准化后的交易记录DataFrame（9列标准字段）
        summary: 汇总信息字典
        output_path: 输出文件路径
        
    返回:
        输出文件路径
        
    异常:
        FileLockedError: 文件被占用时抛出
    """
    logger.info(f"开始导出Excel文件: {output_path}")
    
    try:
        # 使用ExcelWriter创建Excel文件
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Sheet1: Transactions（交易记录）
            df.to_excel(writer, sheet_name='Transactions', index=False)
            
            # Sheet2: Summary（汇总信息）
            summary_df = _create_summary_dataframe(summary)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # 格式化Excel文件
        _format_excel_file(output_path)
        
        logger.info(f"✅ Excel文件导出成功: {output_path}")
        return output_path
        
    except PermissionError as e:
        # Windows文件占用锁
        error_msg = f"⚠️ 文件被占用，请先关闭已打开的 Excel 文件：\n{output_path}"
        logger.error(error_msg)
        raise FileLockedError(error_msg) from e
    except Exception as e:
        logger.error(f"导出Excel文件失败: {e}")
        raise


def _create_summary_dataframe(summary: Dict[str, Any]) -> pd.DataFrame:
    """
    创建汇总信息DataFrame
    
    参数:
        summary: 汇总信息字典
        
    返回:
        汇总信息DataFrame（2列：项目、值）
    """
    # 标准字段顺序
    summary_items = [
        ('原始文件', summary.get('原始文件', '')),
        ('银行', summary.get('银行', '')),
        ('账户币种', summary.get('账户币种', '')),
        ('统计期间', summary.get('统计期间', '')),
        ('期初余额', summary.get('期初余额')),
        ('期末余额', summary.get('期末余额')),
        ('总收入(Credit)', summary.get('总收入(Credit)')),
        ('总支出(Debit)', summary.get('总支出(Debit)')),
        ('交易笔数', summary.get('交易笔数', 0))
    ]
    
    # 构建DataFrame
    data = []
    for item, value in summary_items:
        # 格式化数值
        if value is None:
            display_value = ''
        elif isinstance(value, (int, float)):
            if item in ['期初余额', '期末余额', '总收入(Credit)', '总支出(Debit)']:
                # 金额格式：保留2位小数，显示千位分隔符
                display_value = f"{value:,.2f}"
            else:
                # 交易笔数：整数格式
                display_value = int(value)
        else:
            display_value = str(value)
        
        data.append({
            '项目': item,
            '值': display_value
        })
    
    return pd.DataFrame(data)


def _format_excel_file(file_path: str):
    """
    格式化Excel文件（设置样式、列宽等）
    
    参数:
        file_path: Excel文件路径
    """
    logger.info("开始格式化Excel文件...")
    
    try:
        workbook = load_workbook(file_path)
        
        # 格式化 Transactions Sheet
        if 'Transactions' in workbook.sheetnames:
            _format_transactions_sheet(workbook['Transactions'])
        
        # 格式化 Summary Sheet
        if 'Summary' in workbook.sheetnames:
            _format_summary_sheet(workbook['Summary'])
        
        workbook.save(file_path)
        logger.info("✅ Excel文件格式化完成")
        
    except Exception as e:
        logger.warning(f"格式化Excel文件时出错（文件已创建，但格式可能不完整）: {e}")


def _format_transactions_sheet(worksheet):
    """格式化交易记录Sheet"""
    # 定义样式
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 格式化表头（第1行）
    for col_idx, col_name in enumerate(['Date', 'Account Currency', 'Payer', 'Payee', 
                                        'Debit', 'Credit', 'Balance', 'Reference', 'Description'], 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # 格式化数据行
    for row_idx in range(2, worksheet.max_row + 1):
        for col_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.border = border
            
            # 根据列类型设置格式
            col_letter = get_column_letter(col_idx)
            header_cell = worksheet.cell(row=1, column=col_idx)
            header_value = header_cell.value
            
            if header_value in ['Debit', 'Credit', 'Balance']:
                # 金额列：数值格式，2位小数，千位分隔符
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
                # 确保空值显示为空（不显示0）
                if cell.value is None or cell.value == '':
                    cell.value = None
            elif header_value == 'Date':
                # 日期列：日期格式
                cell.number_format = 'YYYY-MM-DD'
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                # 文本列：左对齐
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # 设置列宽
    column_widths = {
        'A': 12,  # Date
        'B': 15,  # Account Currency
        'C': 25,  # Payer
        'D': 25,  # Payee
        'E': 15,  # Debit
        'F': 15,  # Credit
        'G': 15,  # Balance
        'H': 20,  # Reference
        'I': 50   # Description
    }
    
    for col_letter, width in column_widths.items():
        worksheet.column_dimensions[col_letter].width = width
    
    # 冻结首行
    worksheet.freeze_panes = 'A2'


def _format_summary_sheet(worksheet):
    """格式化汇总信息Sheet"""
    # 定义样式
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 格式化表头（第1行）
    for col_idx in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # 格式化数据行
    for row_idx in range(2, worksheet.max_row + 1):
        for col_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.border = border
            
            if col_idx == 1:
                # 项目列：左对齐，加粗
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.font = Font(bold=True)
            else:
                # 值列：右对齐（数值）或左对齐（文本）
                cell.alignment = Alignment(horizontal="right", vertical="center")
    
    # 设置列宽
    worksheet.column_dimensions['A'].width = 25  # 项目
    worksheet.column_dimensions['B'].width = 30  # 值
    
    # 冻结首行
    worksheet.freeze_panes = 'A2'



